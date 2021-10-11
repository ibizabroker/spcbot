import discord
from discord.ext import commands
from dotenv import load_dotenv
import os
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
import requests

scope = ['https://www.googleapis.com/auth/spreadsheets', 
         'https://www.googleapis.com/auth/drive', 
         'https://www.googleapis.com/auth/drive.file', 
         'https://spreadsheets.google.com/feeds']

credentials = ServiceAccountCredentials.from_json_keyfile_name("google_keys.json", scope)
gspread_client = gspread.authorize(credentials)

Database = gspread_client.open("EC Database").worksheet("Database")

list_usn = Database.col_values(Database.find("USN").col)
list_name = Database.col_values(Database.find("NAME").col)
list_backlog = Database.col_values(Database.find("Backlog").col)
list_placed = Database.col_values(Database.find("Placement Status").col)

list_usn.pop(0)
list_name.pop(0)
list_placed.pop(0)
list_backlog.pop(0)

list_usn_backlog = []
list_name_backlog = []
list_open_dream = []
list_name_open_dream = []
list_dream = []
list_name_dream = []
list_unplaced = []
list_name_unplaced = []

for i in range(len(list_usn)):
    if list_backlog[i] == "Yes":
        list_usn_backlog.append(list_usn[i])
        list_name_backlog.append(list_name[i])

for i in range(len(list_usn)):
    if list_placed[i] == "Open Dream":
        list_open_dream.append(list_usn[i])
        list_name_open_dream.append(list_name[i])
    if list_placed[i] == "Dream":
        list_dream.append(list_usn[i])
        list_name_dream.append(list_name[i])
    if list_placed[i] == "Unplaced":
        list_unplaced.append(list_usn[i])
        list_name_unplaced.append(list_name[i])

load_dotenv()
token = os.getenv("token")

discord_client = commands.Bot(command_prefix = '-', help_command = None, case_insensitive = True)

@discord_client.event
async def on_ready():
    await discord_client.change_presence(activity = discord.Game('-help'))
    print('Logged in as {0.user}'.format(discord_client))

@discord_client.group(invoke_without_command=True)
async def help(ctx):
    embed = discord.Embed(
        title = 'Help', 
        description = 'List of bot commands', 
        colour = discord.Colour.green()
    )
    embed.add_field(name = 'About Student Database', value = 'backlog, placed, unplaced, opendream, dream', inline = False)
    embed.add_field(name = 'Verify', value = 'verify, cutoff', inline = False)
    embed.set_footer(text = 'For more information on a command try -help <command>, ex: -help placed')
    await ctx.send(embed = embed)

@help.command()
async def backlog(ctx):
    embed = discord.Embed(
        title = "Backlog", 
        description = 'Returns a list of USNs of students with backlog', 
        colour = discord.Colour.green()
    )
    embed.add_field(name = '**Syntax**', value = '-backlog')
    await ctx.send(embed = embed)

@help.command()
async def placed(ctx):
    embed = discord.Embed(
        title = "Placed", 
        description = 'Returns a list of USNs of placed students', 
        colour = discord.Colour.blue()
    )
    embed.add_field(name = '**Syntax**', value = '-placed')
    await ctx.send(embed = embed)

@help.command()
async def unplaced(ctx):
    embed = discord.Embed(
        title = "Unplaced", 
        description = 'Returns a list of USNs of unplaced students', 
        colour = discord.Colour.red()
    )
    embed.add_field(name = '**Syntax**', value = '-unplaced')
    await ctx.send(embed = embed)

@help.command()
async def opendream(ctx):
    embed = discord.Embed(
        title = "Open Dream", 
        description = 'Returns a list of USNs of students placed in Open Dream companies', 
        colour = discord.Colour.blue()
    )
    embed.add_field(name = '**Syntax**', value = '-opendream')
    await ctx.send(embed = embed)

@help.command()
async def dream(ctx):
    embed = discord.Embed(
        title = "Dream", 
        description = 'Returns a list of USNs of students placed in Dream companies', 
        colour = discord.Colour.purple()
    )
    embed.add_field(name = '**Syntax**', value = '-dream')
    await ctx.send(embed = embed)

@help.command()
async def verify(ctx):
    embed = discord.Embed(
        title = "Verify", 
        description = 'Returns a list of USNs of students who are already placed in a open dream or a dream company. Also returns a list of USNs of students who have backlogs and have applied for the given company.', 
        colour = discord.Colour.green()
    )
    embed.add_field(name = '**Syntax**', value = '-verify <attachment> Ex: -verify (attach a .xlsx file)')
    await ctx.send(embed = embed)

@help.command()
async def cutoff(ctx):
    embed = discord.Embed(
        title = "Cutoff", 
        description = '', 
        colour = discord.Colour.green()
    )
    embed.add_field(name = '**Syntax**', value = '-cutoff <cgpa> Ex: -cutoff 7.5')
    await ctx.send(embed = embed)

@discord_client.command(aliases = ['hi', 'hey'])
async def hello(ctx):
    await ctx.send('Haaai ;)')

def predicate(message, l, r):
    def check(reaction, user):
        if reaction.message.id != message.id or user == discord_client.user:
            return False
        if l and reaction.emoji == "◀":
            return True
        if r and reaction.emoji == "▶":
            return True
        return False
    return check

@discord_client.command()
async def backlog(ctx):
    # embed = discord.Embed(
    #     title = "Backlog Students",
    #     description = '',
    #     colour = discord.Colour.red() 
    # )

    # # string_usn_backlog = ''
    # # string_name_backlog = ''
    # # for i in range (len(list_usn_backlog)):
    # #     string_usn_backlog += list_usn_backlog[i] + '\n'
    # #     string_name_backlog += list_name_backlog[i] + '\n' 

    # # embed.add_field(
    # #     name = 'USN - Name',
    # #     value = string_usn_backlog + '-' + string_name_backlog,
    # #     inline = False
    # # )
    # for i in range (len(list_usn_backlog)):
    #     embed.add_field(
    #         name = '\u200b',
    #         value = '{} - {}'.format(list_usn_backlog[i], list_name_backlog[i]),
    #         # string_usn_backlog + '-' + string_name_backlog,
    #         inline = False
    #     )
    
    # embed.set_footer(text = 'count of students: {}'.format(len(list_name_backlog)))
    # await ctx.send(embed = embed)
    # await ctx.send(list_usn_backlog)
    
    combined_list_backlog = map(lambda x: x[0]+ ' - ' + x[1], zip(list_usn_backlog, list_name_backlog))
    combined_list_backlog = list(combined_list_backlog)
    composite_list = [combined_list_backlog[x:x + 20] for x in range(0, len(combined_list_backlog), 20)]
    pages = []
    i = 0

    for elements in composite_list:
        string = ""
        for element in elements:
            string += element + "\n"
        embed = discord.Embed(
            title = "Students with backlog",
            colour = discord.Colour.red()
        )
        embed.add_field(name = f"Students list {i+1} - {i + len(elements)} out of {len(combined_list_backlog)}", value = string)
        i += len(elements)
        pages.append(embed)

    page = 0
    left = "◀"
    right = "▶"
    while True:
        msg = await ctx.send(embed = pages[(page)])
        l = page != 0
        r = page != len(pages) - 1
        if l:
            await msg.add_reaction(left)

        if r:
            await msg.add_reaction(right)

        react = await discord_client.wait_for('reaction_add', check=predicate(msg, l, r))
        
        if str(react[0]) == left:
            page -= 1
        elif str(react[0]) == right:
            page += 1

        await msg.delete()
    
@discord_client.command()
async def opendream(ctx):
    combined_list_backlog = map(lambda x: x[0]+ ' - ' + x[1], zip(list_open_dream, list_name_open_dream))
    combined_list_backlog = list(combined_list_backlog)
    composite_list = [combined_list_backlog[x:x + 20] for x in range(0, len(combined_list_backlog), 20)]
    pages = []
    i = 0

    for elements in composite_list:
        string = ""
        for element in elements:
            string += element + "\n"
        embed = discord.Embed(
            title = "Students placed in open dream companies",
            colour = discord.Colour.green()
        )
        embed.add_field(name = f"Students list {i+1} - {i + len(elements)} out of {len(combined_list_backlog)}", value = string)
        i += len(elements)
        pages.append(embed)

    page = 0
    left = "◀"
    right = "▶"
    while True:
        msg = await ctx.send(embed = pages[(page)])
        l = page != 0
        r = page != len(pages) - 1
        if l:
            await msg.add_reaction(left)

        if r:
            await msg.add_reaction(right)

        react = await discord_client.wait_for('reaction_add', check=predicate(msg, l, r))
        
        if str(react[0]) == left:
            page -= 1
        elif str(react[0]) == right:
            page += 1

        await msg.delete()

@discord_client.command()
async def dream(ctx):
    combined_list_backlog = map(lambda x: x[0]+ ' - ' + x[1], zip(list_dream, list_name_dream))
    combined_list_backlog = list(combined_list_backlog)
    composite_list = [combined_list_backlog[x:x + 20] for x in range(0, len(combined_list_backlog), 20)]
    pages = []
    i = 0

    for elements in composite_list:
        string = ""
        for element in elements:
            string += element + "\n"
        embed = discord.Embed(
            title = "Students placed in dream companies",
            colour = discord.Colour.purple()
        )
        embed.add_field(name = f"Students list {i+1} - {i + len(elements)} out of {len(combined_list_backlog)}", value = string)
        i += len(elements)
        pages.append(embed)

    page = 0
    left = "◀"
    right = "▶"
    while True:
        msg = await ctx.send(embed = pages[(page)])
        l = page != 0
        r = page != len(pages) - 1
        if l:
            await msg.add_reaction(left)

        if r:
            await msg.add_reaction(right)

        react = await discord_client.wait_for('reaction_add', check=predicate(msg, l, r))
        
        if str(react[0]) == left:
            page -= 1
        elif str(react[0]) == right:
            page += 1

        await msg.delete()
   
@discord_client.command() 
async def unplaced(ctx):
    combined_list_backlog = map(lambda x: x[0]+ ' - ' + x[1], zip(list_unplaced, list_name_unplaced))
    combined_list_backlog = list(combined_list_backlog)
    composite_list = [combined_list_backlog[x:x + 20] for x in range(0, len(combined_list_backlog), 20)]
    pages = []
    i = 0

    for elements in composite_list:
        string = ""
        for element in elements:
            string += element + "\n"
        embed = discord.Embed(
            title = "Unplaced Students",
            colour = discord.Colour.red()
        )
        embed.add_field(name = f"Students list {i+1} - {i + len(elements)} out of {len(combined_list_backlog)}", value = string)
        i += len(elements)
        pages.append(embed)

    page = 0
    left = "◀"
    right = "▶"
    while True:
        msg = await ctx.send(embed = pages[(page)])
        l = page != 0
        r = page != len(pages) - 1
        if l:
            await msg.add_reaction(left)

        if r:
            await msg.add_reaction(right)

        react = await discord_client.wait_for('reaction_add', check=predicate(msg, l, r))
        
        if str(react[0]) == left:
            page -= 1
        elif str(react[0]) == right:
            page += 1

        await msg.delete()

@discord_client.command()
async def placed(ctx):
    list_placed_usn = list_open_dream + list_dream
    list_name_placed = list_name_open_dream + list_name_dream

    combined_list_backlog = map(lambda x: x[0]+ ' - ' + x[1], zip(list_placed_usn, list_name_placed))
    combined_list_backlog = list(combined_list_backlog)
    composite_list = [combined_list_backlog[x:x + 20] for x in range(0, len(combined_list_backlog), 20)]
    pages = []
    i = 0

    for elements in composite_list:
        string = ""
        for element in elements:
            string += element + "\n"
        embed = discord.Embed(
            title = "Placed Students",
            colour = discord.Colour.blue()
        )
        embed.add_field(name = f"Students list {i+1} - {i + len(elements)} out of {len(combined_list_backlog)}", value = string)
        i += len(elements)
        pages.append(embed)

    page = 0
    left = "◀"
    right = "▶"
    while True:
        msg = await ctx.send(embed = pages[(page)])
        l = page != 0
        r = page != len(pages) - 1
        if l:
            await msg.add_reaction(left)

        if r:
            await msg.add_reaction(right)

        react = await discord_client.wait_for('reaction_add', check=predicate(msg, l, r))
        
        if str(react[0]) == left:
            page -= 1
        elif str(react[0]) == right:
            page += 1

        await msg.delete()

@discord_client.command()
async def verify(ctx):
    await ctx.send('Attach the file you want to verify.')

    # def check(message):
    #     return message.author == ctx.author and bool(message.attachments)
    # await ctx.send('Attach the file you want to verify.')
    # msg = await discord_client.wait_for('message', check=check)
    # attachment_url = msg.attachments[0].url

    def check(message):
        attachments = message.attachments
        if len(attachments) == 0:
            return False
        attachment = attachments[0]
        return attachment.filename.endswith('.xlsx')

    msg = await discord_client.wait_for('message', check=check)
    attachment_url = msg.attachments[0].url

    r = requests.get(attachment_url, allow_redirects=True)
    open('verify.xlsx', 'wb').write(r.content)

    workbook = openpyxl.load_workbook('verify.xlsx')
    worksheet = workbook.active
    list_verify_usn = []
    max_col = worksheet.max_column
    max_row = worksheet.max_row
    for i in range(1, max_col + 1):
        cell = worksheet.cell(row = 1, column = i)
        if cell.value == "USN":
            col_value = i
            break      
    for i in range(2, max_row + 1):
        list_verify_usn.append(worksheet.cell(row = i, column = col_value).value)
    
    list_remove_backlog = [x for x in list_usn_backlog if x in list_verify_usn]
    list_already_opendream = [x for x in list_open_dream if x in list_verify_usn]
    list_already_dream = [x for x in list_dream if x in list_verify_usn] 
    await ctx.send('Already placed in Open Dream: {}'.format(list_already_opendream))
    await ctx.send('Already placed in Dream: {}'.format(list_already_dream))
    await ctx.send('Students with backlog: {}'.format(list_remove_backlog))

# @discord_client.event
# async def on_command_error(ctx, error):
#     if isinstance(error, commands.MissingRequiredArgument):
#         await ctx.send('Wrong usage of command.')    

@discord_client.command()
async def cutoff(ctx, amount: float):
    list_cutoff = []
    list_temp_dream = []
    list_cgpa = Database.col_values(Database.find("CGPA").col)
    list_cgpa.pop(0)
    for i in range(len(list_usn)):
        if (list_cgpa[i] == ''):
            continue
        if float(list_cgpa[i]) > amount and list_placed[i] == 'Unplaced' and list_backlog[i] != 'Yes':
            list_cutoff.append(list_usn[i])
        if float(list_cgpa[i]) > amount and list_placed[i] == 'Dream' and list_backlog[i] != 'Yes':
            list_temp_dream.append(list_usn[i])
    await ctx.send('List of students that can apply:\n{}\nTotal number of eligible students are: {}\n\nList of students who are in dream company and are eligible: {}\nNumber of eligible students placed in Dream company are: {}'.format(list_cutoff, len(list_cutoff), list_temp_dream, len(list_temp_dream)))

discord_client.run(token)